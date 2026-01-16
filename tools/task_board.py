# -*- coding: utf-8 -*-
"""
task_board.py - Excel任务板（ops/tasks.xlsx）管理工具
依赖：openpyxl
用途：
- 管理任务（Tasks Sheet）
- 追加日志（Logs Sheet）
- 导出 PROGRESS.md / TASKS.md
- 分支命名建议 / 从当前git分支推断TASK_ID

推荐（Windows终端不乱码）：
PowerShell:
  [Console]::OutputEncoding = [System.Text.UTF8Encoding]::new()
  $env:PYTHONIOENCODING="utf-8"
CMD:
  chcp 65001
  set PYTHONIOENCODING=utf-8
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import subprocess
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# -----------------------------
# 常量与枚举
# -----------------------------
DEFAULT_XLSX = "ops/tasks.xlsx"
SHEET_TASKS = "Tasks"
SHEET_LOGS = "Logs"
SHEET_ENUMS = "Enums"

STATUS_ENUM = ["TODO", "DOING", "REVIEW", "BLOCKED", "DONE", "RELEASED"]
PRIORITY_ENUM = ["P0", "P1", "P2", "P3"]
ROLE_ENUM = ["PM", "Dev", "Algo", "UI", "QA", "Doc", "Architect", "RDLead"]
BRANCH_TYPE_ENUM = ["feature", "bugfix", "docs"]

# Tasks 表头（严格固定）
TASK_HEADERS = [
    "TASK_ID",
    "需求ID",
    "模块",
    "标题",
    "任务说明",
    "优先级",
    "状态",
    "负责人",
    "角色",
    "分支",
    "进度(0-100)",
    "预估人日",
    "计划开始",
    "计划结束",
    "实际开始",
    "实际结束",
    "依赖任务",
    "验收标准(DoD)",
    "关联文档",
    "测试用例",
    "MR/PR链接",
    "阻塞/风险",
    "最后更新",
    "工作日志",
]

# Logs 表头
LOG_HEADERS = [
    "时间",
    "TASK_ID",
    "动作",
    "内容",
    "分支",
    "CommitHash",
    "执行人",
]

# Enums 表头（简单列表即可）
ENUM_HEADERS = ["类型", "枚举值"]

# 允许 update 的字段（白名单）
UPDATE_ALLOWED_FIELDS = {
    "Status": "状态",
    "Progress": "进度(0-100)",
    "Branch": "分支",
    "Owner": "负责人",
    "Role": "角色",
    "Priority": "优先级",
    "PlanStart": "计划开始",
    "PlanEnd": "计划结束",
    "ActualStart": "实际开始",
    "ActualEnd": "实际结束",
    "ReqId": "需求ID",
    "Module": "模块",
    "Title": "标题",
    "Desc": "任务说明",
    "Deps": "依赖任务",
    "DoD": "验收标准(DoD)",
    "Docs": "关联文档",
    "Cases": "测试用例",
    "MR": "MR/PR链接",
    "Risk": "阻塞/风险",
}

TASK_ID_REGEX = re.compile(r"(PROJ-\d{8}-\d{3})")


# -----------------------------
# 编码：保证终端输出 UTF-8
# -----------------------------
def force_utf8_stdout() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        # 兼容旧版本 Python 或特殊环境
        pass


# -----------------------------
# 工具函数
# -----------------------------
def now_str() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M")


def today_ymd() -> str:
    return dt.datetime.now().strftime("%Y%m%d")


def safe_str(v: Any) -> str:
    """Excel读取出来可能是 datetime 或 None"""
    if v is None:
        return ""
    if isinstance(v, (dt.date, dt.datetime)):
        # 统一格式
        if isinstance(v, dt.datetime):
            return v.strftime("%Y-%m-%d %H:%M")
        return v.strftime("%Y-%m-%d")
    return str(v)


def run_cmd(cmd: List[str], cwd: Optional[str] = None) -> Tuple[int, str, str]:
    """运行外部命令（git），返回 rc, stdout, stderr"""
    try:
        proc = subprocess.run(
            cmd,
            cwd=cwd,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        return proc.returncode, proc.stdout.strip(), proc.stderr.strip()
    except Exception as e:
        return 1, "", f"执行失败: {e}"


def get_git_branch() -> str:
    rc, out, _ = run_cmd(["git", "branch", "--show-current"])
    return out if rc == 0 else ""


def infer_task_id_from_branch(branch: str) -> str:
    m = TASK_ID_REGEX.search(branch or "")
    return m.group(1) if m else ""


def slugify_title(title: str) -> str:
    """
    不引入第三方库，做一个“安全 slug”：
    - 提取 a-zA-Z0-9
    - 空则返回 'task'
    """
    if not title:
        return "task"
    s = title.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = s.strip("-")
    return s if s else "task"


def md_table(rows: List[Dict[str, Any]], cols: List[str], limit: int = 0) -> str:
    """输出 Markdown 表格（避免复杂字符，降低乱码风险）"""
    if limit > 0:
        rows = rows[:limit]

    header = "| " + " | ".join(cols) + " |"
    sep = "| " + " | ".join(["---"] * len(cols)) + " |"
    body_lines = []
    for r in rows:
        body_lines.append("| " + " | ".join(safe_str(r.get(c, "")) for c in cols) + " |")
    return "\n".join([header, sep] + body_lines)


def ensure_parent_dir(path: str) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)


# -----------------------------
# Excel 读写层
# -----------------------------
class TaskBoard:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb: Optional[Workbook] = None

    def load_or_create(self) -> None:
        fp = Path(self.file_path)
        fp.parent.mkdir(parents=True, exist_ok=True)
        if fp.exists():
            self.wb = load_workbook(fp)
        else:
            self.wb = Workbook()
            # 默认会生成一个 "Sheet"，后面我们会清理并创建规范Sheet
            self._init_sheets()

    def save(self) -> None:
        assert self.wb is not None
        try:
            self.wb.save(self.file_path)
        except PermissionError:
            raise PermissionError(f"无法写入Excel：{self.file_path}（可能被Excel占用，请先关闭文件）")

    def _get_sheet(self, name: str) -> Worksheet:
        assert self.wb is not None
        if name in self.wb.sheetnames:
            return self.wb[name]
        return self.wb.create_sheet(name)

    def _clear_default_sheet(self) -> None:
        assert self.wb is not None
        if "Sheet" in self.wb.sheetnames and len(self.wb.sheetnames) > 1:
            ws = self.wb["Sheet"]
            self.wb.remove(ws)

    def _write_headers(self, ws: Worksheet, headers: List[str]) -> None:
        ws.delete_rows(1, ws.max_row)
        ws.append(headers)

    def _init_sheets(self) -> None:
        assert self.wb is not None
        ws_tasks = self._get_sheet(SHEET_TASKS)
        ws_logs = self._get_sheet(SHEET_LOGS)
        ws_enums = self._get_sheet(SHEET_ENUMS)

        self._write_headers(ws_tasks, TASK_HEADERS)
        self._write_headers(ws_logs, LOG_HEADERS)
        self._write_headers(ws_enums, ENUM_HEADERS)

        # Enums 内容
        ws_enums.append(["Status", ",".join(STATUS_ENUM)])
        ws_enums.append(["Priority", ",".join(PRIORITY_ENUM)])
        ws_enums.append(["Role", ",".join(ROLE_ENUM)])
        ws_enums.append(["BranchType", ",".join(BRANCH_TYPE_ENUM)])

        self._clear_default_sheet()
        self.save()

    def init(self) -> None:
        self.load_or_create()
        assert self.wb is not None
        # 强制覆盖为规范sheet
        self._init_sheets()

    # ------- Tasks Sheet 操作 -------
    def tasks_ws(self) -> Worksheet:
        assert self.wb is not None
        return self._get_sheet(SHEET_TASKS)

    def logs_ws(self) -> Worksheet:
        assert self.wb is not None
        return self._get_sheet(SHEET_LOGS)

    def _headers_index(self, ws: Worksheet) -> Dict[str, int]:
        headers = [safe_str(c.value) for c in ws[1]]
        return {h: i for i, h in enumerate(headers)}

    def _iter_rows(self, ws: Worksheet) -> List[Dict[str, Any]]:
        idx = self._headers_index(ws)
        rows: List[Dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            item = {}
            for h, i in idx.items():
                item[h] = r[i] if i < len(r) else None
            rows.append(item)
        return rows

    def list_tasks(self) -> List[Dict[str, Any]]:
        ws = self.tasks_ws()
        return self._iter_rows(ws)

    def find_task_row(self, task_id: str) -> Tuple[int, Dict[str, Any]]:
        """
        返回：(Excel行号, 任务dict)
        Excel行号从1开始，1是表头，因此数据从2开始
        """
        ws = self.tasks_ws()
        idx = self._headers_index(ws)
        tid_col = idx.get("TASK_ID")
        if tid_col is None:
            raise RuntimeError("Tasks 表缺少 TASK_ID 列")

        for row_idx in range(2, ws.max_row + 1):
            cell_val = safe_str(ws.cell(row=row_idx, column=tid_col + 1).value)
            if cell_val == task_id:
                # 读整行
                row_values = {}
                for h, i in idx.items():
                    row_values[h] = ws.cell(row=row_idx, column=i + 1).value
                return row_idx, row_values
        raise KeyError(f"任务不存在：{task_id}")

    def next_task_id(self, project_id: str) -> str:
        """
        生成下一个 TASK_ID：<PROJECT_ID>-NNN
        """
        tasks = self.list_tasks()
        max_seq = 0
        prefix = project_id + "-"
        for t in tasks:
            tid = safe_str(t.get("TASK_ID"))
            if tid.startswith(prefix):
                m = re.search(r"-(\d{3})$", tid)
                if m:
                    max_seq = max(max_seq, int(m.group(1)))
        return f"{project_id}-{max_seq+1:03d}"

    def create_task(
            self,
            project_id: str,
            title: str,
            priority: str,
            role: str,
            owner: str,
            req_id: str,
            module: str = "",
            desc: str = "",
    ) -> Dict[str, Any]:
        ws = self.tasks_ws()
        task_id = self.next_task_id(project_id)

        if priority not in PRIORITY_ENUM:
            raise ValueError(f"priority 必须为 {PRIORITY_ENUM}")
        if role not in ROLE_ENUM:
            raise ValueError(f"role 必须为 {ROLE_ENUM}")

        row = {
            "TASK_ID": task_id,
            "需求ID": req_id,
            "模块": module,
            "标题": title,
            "任务说明": desc,
            "优先级": priority,
            "状态": "TODO",
            "负责人": owner,
            "角色": role,
            "分支": "",
            "进度(0-100)": 0,
            "预估人日": "",
            "计划开始": "",
            "计划结束": "",
            "实际开始": "",
            "实际结束": "",
            "依赖任务": "",
            "验收标准(DoD)": "",
            "关联文档": "",
            "测试用例": "",
            "MR/PR链接": "",
            "阻塞/风险": "",
            "最后更新": now_str(),
            "工作日志": "",
        }

        idx = self._headers_index(ws)
        # 按表头顺序 append
        ws.append([row.get(h, "") for h in TASK_HEADERS])
        self.save()
        return row

    def suggest_branch(self, task_id: str) -> str:
        _, task = self.find_task_row(task_id)
        title = safe_str(task.get("标题"))
        role = safe_str(task.get("角色"))

        slug = slugify_title(title)

        if role == "Doc":
            return f"docs/{task_id}-{slug}"
        if role in ("QA",):
            return f"bugfix/{task_id}-{slug}"
        # 默认 feature
        return f"feature/{task_id}-{slug}"

    def update_task(self, task_id: str, kvs: Dict[str, str]) -> Dict[str, Any]:
        ws = self.tasks_ws()
        row_idx, current = self.find_task_row(task_id)
        idx = self._headers_index(ws)

        # 校验 + 更新
        for k, v in kvs.items():
            if k not in UPDATE_ALLOWED_FIELDS:
                raise ValueError(f"不允许更新字段：{k}，允许：{list(UPDATE_ALLOWED_FIELDS.keys())}")
            col_name = UPDATE_ALLOWED_FIELDS[k]
            if col_name not in idx:
                raise RuntimeError(f"Tasks 表缺少列：{col_name}")

            # 校验枚举字段
            if col_name == "状态":
                if v not in STATUS_ENUM:
                    raise ValueError(f"Status 必须为 {STATUS_ENUM}")
            if col_name == "优先级":
                if v not in PRIORITY_ENUM:
                    raise ValueError(f"Priority 必须为 {PRIORITY_ENUM}")
            if col_name == "角色":
                if v not in ROLE_ENUM:
                    raise ValueError(f"Role 必须为 {ROLE_ENUM}")
            if col_name == "进度(0-100)":
                try:
                    pv = int(v)
                except ValueError:
                    raise ValueError("Progress 必须是整数 0~100")
                if pv < 0 or pv > 100:
                    raise ValueError("Progress 超出范围 0~100")
                v = pv  # 写入数字

            # DONE 规则：进度建议 >= 95
            if col_name == "状态" and v == "DONE":
                prog = current.get("进度(0-100)")
                try:
                    prog_int = int(prog) if prog is not None and safe_str(prog) != "" else 0
                except ValueError:
                    prog_int = 0
                if prog_int < 95:
                    raise ValueError("状态设为 DONE 前，进度建议 >= 95（请先更新 Progress）")

            ws.cell(row=row_idx, column=idx[col_name] + 1).value = v
            current[col_name] = v

        # 最后更新
        if "最后更新" in idx:
            ws.cell(row=row_idx, column=idx["最后更新"] + 1).value = now_str()
            current["最后更新"] = now_str()

        self.save()
        return {h: current.get(h, "") for h in TASK_HEADERS}

    def append_log(
            self,
            task_id: str,
            content: str,
            action: str = "update",
            branch: str = "",
            commit_hash: str = "",
            actor: str = "",
    ) -> None:
        ws_tasks = self.tasks_ws()
        ws_logs = self.logs_ws()
        row_idx, current = self.find_task_row(task_id)
        idx = self._headers_index(ws_tasks)

        if not branch:
            branch = get_git_branch()

        line = f"[{now_str()}] {content}".strip()

        # 更新 Tasks.WorkLog（多行追加）
        worklog_col = idx.get("工作日志")
        if worklog_col is not None:
            old = safe_str(ws_tasks.cell(row=row_idx, column=worklog_col + 1).value)
            new_val = (old + "\n" + line).strip() if old else line
            ws_tasks.cell(row=row_idx, column=worklog_col + 1).value = new_val

        # 更新最后更新
        last_col = idx.get("最后更新")
        if last_col is not None:
            ws_tasks.cell(row=row_idx, column=last_col + 1).value = now_str()

        # 写 Logs Sheet
        ws_logs.append([now_str(), task_id, action, content, branch, commit_hash, actor])

        self.save()

    # ------- 导出 Markdown -------
    def export_md(self, progress_path: str, tasks_path: str) -> None:
        tasks = self.list_tasks()

        # PROGRESS.md 字段
        progress_cols = [
            "TASK_ID",
            "需求ID",
            "标题",
            "优先级",
            "状态",
            "负责人",
            "角色",
            "分支",
            "进度(0-100)",
            "计划开始",
            "计划结束",
            "实际开始",
            "实际结束",
            "阻塞/风险",
            "最后更新",
            "MR/PR链接",
            "备注",
        ]

        # TASKS.md 字段（更详细）
        tasks_cols = [
            "TASK_ID",
            "需求ID",
            "模块",
            "标题",
            "任务说明",
            "Owner角色",
            "Owner",
            "优先级",
            "状态",
            "预估(人日)",
            "分支(branch)",
            "依赖任务",
            "验收标准(DoD)",
            "关联文档",
            "测试用例",
            "MR/PR链接",
            "最后更新",
            "备注",
        ]

        # 组装行（映射字段名）
        prog_rows = []
        task_rows = []

        # 排序：P0优先 + 状态 DOING/TODO/REVIEW/BLOCKED/DONE/RELEASED
        status_rank = {s: i for i, s in enumerate(["DOING", "TODO", "REVIEW", "BLOCKED", "DONE", "RELEASED"])}
        pr_rank = {p: i for i, p in enumerate(PRIORITY_ENUM)}

        def sort_key(t: Dict[str, Any]):
            return (
                pr_rank.get(safe_str(t.get("优先级")), 99),
                status_rank.get(safe_str(t.get("状态")), 99),
                safe_str(t.get("TASK_ID")),
            )

        tasks_sorted = sorted(tasks, key=sort_key)

        for t in tasks_sorted:
            prog_rows.append(
                {
                    "TASK_ID": safe_str(t.get("TASK_ID")),
                    "需求ID": safe_str(t.get("需求ID")),
                    "标题": safe_str(t.get("标题")),
                    "优先级": safe_str(t.get("优先级")),
                    "状态": safe_str(t.get("状态")),
                    "负责人": safe_str(t.get("负责人")),
                    "角色": safe_str(t.get("角色")),
                    "分支": safe_str(t.get("分支")),
                    "进度(0-100)": safe_str(t.get("进度(0-100)")),
                    "计划开始": safe_str(t.get("计划开始")),
                    "计划结束": safe_str(t.get("计划结束")),
                    "实际开始": safe_str(t.get("实际开始")),
                    "实际结束": safe_str(t.get("实际结束")),
                    "阻塞/风险": safe_str(t.get("阻塞/风险")),
                    "最后更新": safe_str(t.get("最后更新")),
                    "MR/PR链接": safe_str(t.get("MR/PR链接")),
                    "备注": "",
                }
            )

            task_rows.append(
                {
                    "TASK_ID": safe_str(t.get("TASK_ID")),
                    "需求ID": safe_str(t.get("需求ID")),
                    "模块": safe_str(t.get("模块")),
                    "标题": safe_str(t.get("标题")),
                    "任务说明": safe_str(t.get("任务说明")),
                    "Owner角色": safe_str(t.get("角色")),
                    "Owner": safe_str(t.get("负责人")),
                    "优先级": safe_str(t.get("优先级")),
                    "状态": safe_str(t.get("状态")),
                    "预估(人日)": safe_str(t.get("预估人日")),
                    "分支(branch)": safe_str(t.get("分支")),
                    "依赖任务": safe_str(t.get("依赖任务")),
                    "验收标准(DoD)": safe_str(t.get("验收标准(DoD)")),
                    "关联文档": safe_str(t.get("关联文档")),
                    "测试用例": safe_str(t.get("测试用例")),
                    "MR/PR链接": safe_str(t.get("MR/PR链接")),
                    "最后更新": safe_str(t.get("最后更新")),
                    "备注": "",
                }
            )

        ensure_parent_dir(progress_path)
        ensure_parent_dir(tasks_path)

        progress_md = "# 项目进度表 PROGRESS\n\n" + md_table(prog_rows, progress_cols) + "\n"
        tasks_md = "# 任务清单 TASKS\n\n" + md_table(task_rows, tasks_cols) + "\n"

        with open(progress_path, "w", encoding="utf-8") as f:
            f.write(progress_md)

        with open(tasks_path, "w", encoding="utf-8") as f:
            f.write(tasks_md)


# -----------------------------
# 输出格式（pretty / json）
# -----------------------------
def output(data: Any, pretty: bool, json_out: bool) -> None:
    if json_out:
        print(json.dumps(data, ensure_ascii=False, indent=2))
        return
    if pretty:
        # pretty 优先使用 Markdown
        if isinstance(data, list):
            if not data:
                print("(空)")
                return
            # 自动选 cols
            cols = list(data[0].keys())
            print(md_table(data, cols))
        elif isinstance(data, dict):
            # dict 输出为 key: value 列表
            for k, v in data.items():
                print(f"- {k}: {safe_str(v)}")
        else:
            print(safe_str(data))
    else:
        print(safe_str(data))


# -----------------------------
# CLI 子命令
# -----------------------------
def cmd_init(args: argparse.Namespace) -> None:
    tb = TaskBoard(args.file)
    tb.init()
    output({"result": "ok", "file": args.file, "sheets": [SHEET_TASKS, SHEET_LOGS, SHEET_ENUMS]}, args.pretty, args.json)


def cmd_list(args: argparse.Namespace) -> None:
    tb = TaskBoard(args.file)
    tb.load_or_create()
    tasks = tb.list_tasks()

    # 过滤
    def match(t: Dict[str, Any]) -> bool:
        if args.status:
            st = safe_str(t.get("状态"))
            if st not in args.status:
                return False
        if args.owner:
            if safe_str(t.get("负责人")) != args.owner:
                return False
        if args.role:
            if safe_str(t.get("角色")) != args.role:
                return False
        if args.priority:
            if safe_str(t.get("优先级")) != args.priority:
                return False
        if args.contains:
            needle = args.contains
            if needle not in safe_str(t.get("标题")) and needle not in safe_str(t.get("任务说明")):
                return False
        return True

    tasks = [t for t in tasks if match(t)]

    # 默认精简字段
    rows = []
    for t in tasks:
        rows.append(
            {
                "TASK_ID": safe_str(t.get("TASK_ID")),
                "标题": safe_str(t.get("标题")),
                "状态": safe_str(t.get("状态")),
                "优先级": safe_str(t.get("优先级")),
                "负责人": safe_str(t.get("负责人")),
                "分支": safe_str(t.get("分支")),
                "进度": safe_str(t.get("进度(0-100)")),
                "最后更新": safe_str(t.get("最后更新")),
            }
        )

    output(rows, args.pretty, args.json)


def cmd_get(args: argparse.Namespace) -> None:
    tb = TaskBoard(args.file)
    tb.load_or_create()
    _, task = tb.find_task_row(args.task_id)
    # 输出全字段
    full = {h: safe_str(task.get(h, "")) for h in TASK_HEADERS}
    output(full, args.pretty, args.json)


def cmd_create(args: argparse.Namespace) -> None:
    tb = TaskBoard(args.file)
    tb.load_or_create()

    project_id = args.project_id or f"PROJ-{today_ymd()}"
    row = tb.create_task(
        project_id=project_id,
        title=args.title,
        priority=args.priority,
        role=args.role,
        owner=args.owner,
        req_id=args.req,
        module=args.module or "",
        desc=args.desc or "",
    )
    output(row, args.pretty, args.json)


def cmd_suggest_branch(args: argparse.Namespace) -> None:
    tb = TaskBoard(args.file)
    tb.load_or_create()
    branch = tb.suggest_branch(args.task_id)
    output({"task_id": args.task_id, "branch": branch}, args.pretty, args.json)


def cmd_update(args: argparse.Namespace) -> None:
    tb = TaskBoard(args.file)
    tb.load_or_create()

    kvs = {}
    for a in args.assign:
        if "=" not in a:
            raise ValueError(f"非法参数：{a}，应为 Key=Value")
        k, v = a.split("=", 1)
        kvs[k] = v

    updated = tb.update_task(args.task_id, kvs)
    output(updated, args.pretty, args.json)


def cmd_append_log(args: argparse.Namespace) -> None:
    tb = TaskBoard(args.file)
    tb.load_or_create()

    branch = args.branch or get_git_branch()
    tb.append_log(
        task_id=args.task_id,
        content=args.message,
        action=args.action or "update",
        branch=branch,
        commit_hash=args.commit or "",
        actor=args.actor or "",
    )
    output({"result": "ok", "task_id": args.task_id}, args.pretty, args.json)


def cmd_infer_task_id(args: argparse.Namespace) -> None:
    branch = get_git_branch()
    task_id = infer_task_id_from_branch(branch)
    output({"branch": branch, "task_id": task_id}, args.pretty, args.json)


def cmd_export_md(args: argparse.Namespace) -> None:
    tb = TaskBoard(args.file)
    tb.load_or_create()
    tb.export_md(args.progress, args.tasks)
    output(
        {"result": "ok", "progress": args.progress, "tasks": args.tasks},
        args.pretty,
        args.json,
    )


# -----------------------------
# main
# -----------------------------
def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="task_board.py", add_help=True)

    p.add_argument("--file", default=DEFAULT_XLSX, help="任务板文件路径（默认 ops/tasks.xlsx）")
    p.add_argument("--pretty", action="store_true", help="Markdown友好输出")
    p.add_argument("--json", action="store_true", help="JSON输出")

    sub = p.add_subparsers(dest="cmd", required=True)

    sp = sub.add_parser("init", help="初始化任务板（创建Tasks/Logs/Enums）")
    sp.set_defaults(func=cmd_init)

    sp = sub.add_parser("list", help="列出任务")
    sp.add_argument("--status", type=str, default="", help="状态过滤：TODO,DOING,...（逗号分隔）")
    sp.add_argument("--owner", type=str, default="", help="负责人过滤")
    sp.add_argument("--role", type=str, default="", help="角色过滤：Dev/UI/QA/Doc...")
    sp.add_argument("--priority", type=str, default="", help="优先级过滤：P0/P1/P2/P3")
    sp.add_argument("--contains", type=str, default="", help="标题/说明模糊匹配关键字")
    sp.set_defaults(func=cmd_list)

    sp = sub.add_parser("get", help="查看单个任务详情")
    sp.add_argument("--task-id", required=True)
    sp.set_defaults(func=cmd_get)

    sp = sub.add_parser("create", help="创建任务（PM专用）")
    sp.add_argument("--project-id", default="", help="项目ID（默认 PROJ-YYYYMMDD）")
    sp.add_argument("--title", required=True, help="任务标题")
    sp.add_argument("--priority", required=True, choices=PRIORITY_ENUM)
    sp.add_argument("--role", required=True, choices=ROLE_ENUM)
    sp.add_argument("--owner", required=True)
    sp.add_argument("--req", required=True, help="需求ID，如 PRD-001")
    sp.add_argument("--module", default="", help="模块名")
    sp.add_argument("--desc", default="", help="任务说明")
    sp.set_defaults(func=cmd_create)

    sp = sub.add_parser("suggest-branch", help="为任务建议分支名")
    sp.add_argument("--task-id", required=True)
    sp.set_defaults(func=cmd_suggest_branch)

    sp = sub.add_parser("update", help="更新任务字段（Key=Value）")
    sp.add_argument("--task-id", required=True)
    sp.add_argument("assign", nargs="+", help="Key=Value ... （例：Status=DOING Progress=20）")
    sp.set_defaults(func=cmd_update)

    sp = sub.add_parser("append-log", help="追加日志（同时写Tasks.WorkLog与Logs表）")
    sp.add_argument("--task-id", required=True)
    sp.add_argument("message", help="日志内容（中文）")
    sp.add_argument("--action", default="update", help="动作：start/commit/finish/switch/update")
    sp.add_argument("--branch", default="", help="分支（默认从git读取）")
    sp.add_argument("--commit", default="", help="commit hash（可选）")
    sp.add_argument("--actor", default="", help="执行人（可选）")
    sp.set_defaults(func=cmd_append_log)

    sp = sub.add_parser("infer-task-id", help="从当前git分支推断TASK_ID")
    sp.set_defaults(func=cmd_infer_task_id)

    sp = sub.add_parser("export-md", help="导出 PROGRESS.md / TASKS.md")
    sp.add_argument("--progress", required=True, help="输出进度表md路径")
    sp.add_argument("--tasks", required=True, help="输出任务表md路径")
    sp.set_defaults(func=cmd_export_md)

    return p


def main() -> int:
    force_utf8_stdout()
    parser = build_parser()
    args = parser.parse_args()

    # 参数清洗
    if hasattr(args, "status") and args.status:
        args.status = [s.strip() for s in args.status.split(",") if s.strip()]
    else:
        if hasattr(args, "status"):
            args.status = []

    if hasattr(args, "owner") and args.owner == "":
        args.owner = ""
    if hasattr(args, "role") and args.role == "":
        args.role = ""
    if hasattr(args, "priority") and args.priority == "":
        args.priority = ""
    if hasattr(args, "contains") and args.contains == "":
        args.contains = ""

    try:
        args.func(args)
        return 0
    except KeyError as e:
        print(str(e), file=sys.stderr)
        return 3
    except PermissionError as e:
        print(str(e), file=sys.stderr)
        return 4
    except ValueError as e:
        print(str(e), file=sys.stderr)
        return 2
    except Exception as e:
        print(f"未知错误：{e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
